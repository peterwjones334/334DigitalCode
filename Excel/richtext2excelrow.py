from spellchecker import SpellChecker
import tkinter as tk
from tkinter import ttk, font, Menu
import openpyxl
from openpyxl import Workbook

# Check if the Excel file already exists to either load it or create a new one
try:
    wb = openpyxl.load_workbook('data.xlsx')
    ws = wb.active
except FileNotFoundError:
    wb = Workbook()
    ws = wb.active
    # Create headers for the Excel sheet
    ws.append(['ID', 'Title', 'Description', 'Priority'])

# Function to submit form data to the Excel sheet
def submit():
    global ws
    last_row = ws.max_row
    new_id = last_row  # Assuming the first row is headers and IDs start from 1

    # Retrieve form values
    title = title_entry.get()
    description = description_entry.get("1.0", "end-1c")  # Text widget
    priority = priority_combobox.get()

    # Append new row to the Excel sheet
    ws.append([new_id, title, description, priority])

    # Clear form fields
    title_entry.delete(0, tk.END)
    description_entry.delete("1.0", tk.END)
    priority_combobox.set('')

# Function to save the workbook and close the form
def save_and_close():
    # Save the workbook
    wb.save('data.xlsx')
    # Close the Tkinter window
    root.destroy()        

# root = tk.Tk()
# root.title("Rich Text Entry Example")

# ttk.Label(root, text="Title:").grid(row=0, column=0, padx=10, pady=5)
# title_entry = ttk.Entry(root)
# title_entry.grid(row=0, column=1, padx=10, pady=5)

# Description field (using Text widget for multi-line input)
# ttk.Label(root, text="Description:").grid(row=1, column=0, padx=10, pady=5)
# description_entry = tk.Text(root, height=5, width=25)
# description_entry.grid(row=1, column=1, padx=10, pady=5)

# description_entry = tk.Text(root, height=20, width=80)
# description_entry.pack(padx=10, pady=10)

# Priority field (ComboBox with values 1-5)
#ttk.Label(root, text="Priority:").grid(row=2, column=0, padx=10, pady=5)
#priority_combobox = ttk.Combobox(root, values=[1, 2, 3, 4, 5], state="readonly")
#priority_combobox.grid(row=2, column=1, padx=10, pady=5)

# Submit button
# submit_button = ttk.Button(root, text="Submit", command=submit)
# submit_button.grid(row=3, column=0, padx=10, pady=10)

# Save & Close button
#save_close_button = ttk.Button(root, text="Save & Close", command=save_and_close)
#save_close_button.grid(row=3, column=1, padx=10, pady=10)

# Initialize the spell checker
spell = SpellChecker()





def set_text_red():
    try:
        description_entry.tag_add("red_text", "sel.first", "sel.last")
    except tk.TclError:
        print("No text selected")

def set_text_black():
    try:
        description_entry.tag_add("black_text", "sel.first", "sel.last")
    except tk.TclError:
        print("No text selected")        

def toggle_bold():
    try:
        current_tags = description_entry.tag_names("sel.first")
        if "bold" in current_tags or "bold_italic" in current_tags:
            description_entry.tag_remove("bold", "sel.first", "sel.last")
            description_entry.tag_remove("bold_italic", "sel.first", "sel.last")
            if "italic" in current_tags:
                description_entry.tag_add("italic", "sel.first", "sel.last")
        else:
            description_entry.tag_add("bold", "sel.first", "sel.last")
            if "italic" in current_tags:
                description_entry.tag_add("bold_italic", "sel.first", "sel.last")
                description_entry.tag_remove("bold", "sel.first", "sel.last")
                description_entry.tag_remove("italic", "sel.first", "sel.last")
    except tk.TclError:
        print("No text selected")

def toggle_italic():
    try:
        current_tags = description_entry.tag_names("sel.first")
        if "italic" in current_tags or "bold_italic" in current_tags:
            description_entry.tag_remove("italic", "sel.first", "sel.last")
            description_entry.tag_remove("bold_italic", "sel.first", "sel.last")
            if "bold" in current_tags:
                description_entry.tag_add("bold", "sel.first", "sel.last")
        else:
            description_entry.tag_add("italic", "sel.first", "sel.last")
            if "bold" in current_tags:
                description_entry.tag_add("bold_italic", "sel.first", "sel.last")
                description_entry.tag_remove("bold", "sel.first", "sel.last")
                description_entry.tag_remove("italic", "sel.first", "sel.last")
    except tk.TclError:
        print("No text selected")

def reset_formatting():
    try:
        description_entry.tag_remove("bold", "sel.first", "sel.last")
        description_entry.tag_remove("italic", "sel.first", "sel.last")
        description_entry.tag_remove("bold_italic", "sel.first", "sel.last")
        description_entry.tag_remove("red_text", "sel.first", "sel.last")
        description_entry.tag_remove("misspelled", "1.0", tk.END)
    except tk.TclError:
        print("No text selected")

def highlight_misspelled():
    description_entry.tag_remove('misspelled', '1.0', tk.END)
    text = description_entry.get('1.0', tk.END)
    words = text.split()

    for word in spell.unknown(words):
        start_index = '1.0'
        while True:
            start_index = description_entry.search(word, start_index, stopindex=tk.END)
            if not start_index:
                break
            end_index = f"{start_index}+{len(word)}c"
            description_entry.tag_add('misspelled', start_index, end_index)
            start_index = end_index

def on_right_click(event):
    try:
        # Clear existing menu
        for item in context_menu.winfo_children():
            item.destroy()

        # Get the position of the cursor
        position = description_entry.index(tk.CURRENT)
        word_start = f"{position} wordstart"
        word_end = f"{position} wordend"

        # Get the word under the cursor
        word = description_entry.get(word_start, word_end)

        # If the word is misspelled, show suggestions
        if "misspelled" in description_entry.tag_names(position):
            suggestions = spell.candidates(word)

            # Check if suggestions is not None and is iterable
            if suggestions:
                for suggestion in suggestions:
                    context_menu.add_command(label=suggestion, command=lambda suggestion=suggestion: replace_word(word_start, word_end, suggestion))
                context_menu.tk_popup(event.x_root, event.y_root)
            else:
                # Handle case with no suggestions
                context_menu.add_command(label="No suggestions", command=lambda: None)
                context_menu.tk_popup(event.x_root, event.y_root)

    finally:
        context_menu.grab_release()

def replace_word(start, end, replacement):
    description_entry.delete(start, end)
    description_entry.insert(start, replacement)
    highlight_misspelled()  # Recheck spelling after replacement        



root = tk.Tk()
root.title("Rich Text Entry Example")

ttk.Label(root, text="Title:").grid(row=0, column=0, padx=5, pady=5)
title_entry = ttk.Entry(root)
title_entry.grid(row=0, column=1, padx=5, pady=5,columnspan=2)

# Priority field (ComboBox with values 1-5)
ttk.Label(root, text="Priority:").grid(row=0, column=3, padx=5, pady=5)
priority_combobox = ttk.Combobox(root, height=5, width=5, values=[1, 2, 3, 4, 5], state="readonly")
priority_combobox.grid(row=0, column=4, padx=5, pady=5)
priority_combobox.set(3)  # Set default value to 3

# Description field (using Text widget for multi-line input)
ttk.Label(root, text="Description:").grid(row=1, column=0, padx=5, pady=5, sticky='e')
description_entry = tk.Text(root, height=10, width=10)
description_entry.grid(row=2, column=0, padx=5, pady=5, sticky='ew', columnspan=5)

# Submit button

reset_button = ttk.Button(root, text="Reset", command=reset_formatting)
reset_button.grid(row=4, column=0, padx=5, pady=5, sticky='ew')

bold_button = ttk.Button(root, text="Bold", command=toggle_bold)
bold_button.grid(row=4, column=1, padx=5, pady=5, sticky='ew')

italic_button = ttk.Button(root, text="Italic", command=toggle_italic)
italic_button.grid(row=4, column=2, padx=5, pady=5, sticky='ew')

red_button = ttk.Button(root, text="Red", command=set_text_red)  # Button to make text red
red_button.grid(row=4, column=3, padx=5, pady=5, sticky='ew')

black_button = ttk.Button(root, text="Black", command=set_text_black)  # Button to make text Black
black_button.grid(row=4, column=4, padx=5, pady=5, sticky='ew')

spell_check_button = ttk.Button(root, text="Spell Check", command=highlight_misspelled)
spell_check_button.grid(row=5, column=0, padx=5, pady=5, sticky='ew')

submit_button = ttk.Button(root, text="Submit", command=submit)
submit_button.grid(row=5, column=1, padx=5, pady=5, sticky='ew')

# Save & Close button
save_close_button = ttk.Button(root, text="Save & Close", command=save_and_close)
save_close_button.grid(row=5, column=2, padx=5, pady=5, sticky='ew')

# Configure tags for styling
normal_font = font.Font(description_entry, description_entry.cget("font"))
bold_font = normal_font.copy()
bold_font.configure(weight="bold")
italic_font = normal_font.copy()
italic_font.configure(slant="italic")
bold_italic_font = normal_font.copy()
bold_italic_font.configure(weight="bold", slant="italic")

description_entry.tag_configure("bold", font=bold_font)
description_entry.tag_configure("italic", font=italic_font)
description_entry.tag_configure("bold_italic", font=bold_italic_font)
description_entry.tag_configure("red_text", foreground="red")  # Red text tag configuration
description_entry.tag_configure("black_text", foreground="black")  # Black text tag configuration
description_entry.tag_configure("misspelled", underline=True, underlinefg='red')

# Create a popup menu for spell correction suggestions
context_menu = Menu(description_entry, tearoff=0)

# Binding the right-click event
description_entry.bind("<Button-3>", on_right_click)

root.mainloop()

