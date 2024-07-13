import tkinter as tk
from tkinter import ttk
import openpyxl
from openpyxl import Workbook

# Check if the Excel file already exists to either load it or create a new one
try:
    wb = openpyxl.load_workbook('data.xlsx')
    ws = wb.active
except FileNotFoundError:
    wb = Workbook()
    ws = wb.active
    # Create headers for the Excel sheet
    ws.append(['ID', 'Title', 'Description', 'Priority'])

# Function to submit form data to the Excel sheet
def submit():
    global ws
    last_row = ws.max_row
    new_id = last_row  # Assuming the first row is headers and IDs start from 1

    # Retrieve form values
    title = title_entry.get()
    description = description_entry.get("1.0", "end-1c")  # Text widget
    priority = priority_combobox.get()

    # Append new row to the Excel sheet
    ws.append([new_id, title, description, priority])

    # Clear form fields
    title_entry.delete(0, tk.END)
    description_entry.delete("1.0", tk.END)
    priority_combobox.set('')

# Function to save the workbook and close the form
def save_and_close():
    # Save the workbook
    wb.save('data.xlsx')
    # Close the Tkinter window
    root.destroy()

# Initialize the Tkinter root window
root = tk.Tk()
root.title("Data Entry Form")

# Title field
ttk.Label(root, text="Title:").grid(row=0, column=0, padx=10, pady=5)
title_entry = ttk.Entry(root)
title_entry.grid(row=0, column=1, padx=10, pady=5)

# Description field (using Text widget for multi-line input)
ttk.Label(root, text="Description:").grid(row=1, column=0, padx=10, pady=5)
description_entry = tk.Text(root, height=5, width=25)
description_entry.grid(row=1, column=1, padx=10, pady=5)

# Priority field (ComboBox with values 1-5)
ttk.Label(root, text="Priority:").grid(row=2, column=0, padx=10, pady=5)
priority_combobox = ttk.Combobox(root, values=[1, 2, 3, 4, 5], state="readonly")
priority_combobox.grid(row=2, column=1, padx=10, pady=5)

# Submit button
submit_button = ttk.Button(root, text="Submit", command=submit)
submit_button.grid(row=3, column=0, padx=10, pady=10)

# Save & Close button
save_close_button = ttk.Button(root, text="Save & Close", command=save_and_close)
save_close_button.grid(row=3, column=1, padx=10, pady=10)

# Run the Tkinter event loop
root.mainloop()

