import openpyxl
from openpyxl import Workbook
import tkinter as tk
from tkinter import ttk
import datetime

root = tk.Tk()
root.title("Requirement Form")

wb = Workbook()
ws = wb.active
ws.title = "Requirements"

# Add 'ID' at the beginning of your header row
ws['A1'] = 'ID'
ws['B1'] = 'Date'
ws['C1'] = 'Name'
ws['D1'] = 'Description'
ws['E1'] = 'Type'
ws['F1'] = 'Priority'
ws['G1'] = 'Dependencies'
ws['H1'] = 'Rationale'
ws['I1'] = 'Source'

def save_workbook():
    # Specify the file path and name for your workbook
    file_path = "requirements.xlsx"
    wb.save(file_path)
    print(f"Workbook saved as '{file_path}'.")

def submit_requirement():

    # Initialize a global variable for ID
    current_id = 1

    now = datetime.datetime.now()  # Get the current time for each submission
    formatted_now = now.strftime("%Y-%m-%d %H:%M:%S")  # Format the timestamp

    name = name_entry.get()
    description = description_entry.get("1.0", "end-1c")
    typ = type_combobox.get()
    priority = priority_combobox.get()
    dependencies = dependencies_entry.get("1.0", "end-1c")
    rationale = rationale_entry.get("1.0", "end-1c")
    source = source_entry.get()

    # Add the current_id at the beginning of your row
    ws.append([current_id, formatted_now, name, description, typ, priority, dependencies, rationale, source])
    current_id += 1  # Increment the ID for the next entry

    name_entry.delete(0, tk.END)  # Corrected deletion method for Entry widgets
    description_entry.delete("1.0", tk.END)
    type_combobox.set('')
    priority_combobox.set('')
    dependencies_entry.delete("1.0", tk.END)
    rationale_entry.delete("1.0", tk.END)
    source_entry.delete(0, tk.END)  # Corrected deletion method for Entry widgets
   
    
name_label = ttk.Label(root, text="Name:")
name_label.grid(row=0, column=0, padx=10, pady=10)
name_entry = ttk.Entry(root, width=30)
name_entry.grid(row=0, column=1, padx=10, pady=10)

description_label = ttk.Label(root, text="Description:")
description_label.grid(row=1, column=0, padx=10, pady=10)
description_entry = tk.Text(root, height=5, width=30)
description_entry.grid(row=1, column=1, padx=10, pady=10)

type_label = ttk.Label(root, text="Type:")
type_label.grid(row=2, column=0, padx=10, pady=10)
type_options = ['Functional', 'Non-Functional', 'Security', 'Data', 'Interface', 'Operational', 'Regulatory']
type_combobox = ttk.Combobox(root, state='readonly', values=type_options)
type_combobox.current(0)
type_combobox.grid(row=2, column=1, padx=10, pady=10)

priority_label = ttk.Label(root, text="Priority:")
priority_label.grid(row=3, column=0, padx=10, pady=10)
priority_options = ['High', 'Medium', 'Low']
priority_combobox = ttk.Combobox(root, state='readonly', values=priority_options)
priority_combobox.current(0)
priority_combobox.grid(row=3, column=1, padx=10, pady=10)

dependencies_label = ttk.Label(root, text="Dependencies:")
dependencies_label.grid(row=4, column=0, padx=10, pady=10)
dependencies_entry = tk.Text(root, height=1, width=30)
dependencies_entry.grid(row=4, column=1, padx=10, pady=10)

rationale_label = ttk.Label(root, text="Rationale:")
rationale_label.grid(row=5, column=0, padx=10, pady=10)
rationale_entry = tk.Text(root, height=5, width=30)
rationale_entry.grid(row=5, column=1, padx=10, pady=10)

source_label = ttk.Label(root, text="Source:")
source_label.grid(row=6, column=0, padx=10, pady=10)
source_entry = ttk.Entry(root, width=30)
source_entry.grid(row=6, column=1, padx=10, pady=10)

submit_button = ttk.Button(root, text="Submit", command=submit_requirement)
submit_button.grid(row=7, column=1, padx=10, pady=10)

# Add the Save button to your form
save_button = ttk.Button(root, text="Save", command=save_workbook)
save_button.grid(row=8, column=1, padx=10, pady=10)  # Adjust row and column as needed
  
root.mainloop()

wb.save("requirements.xlsx")

